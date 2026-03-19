# -*- coding: utf-8 -*-
import os
import json
import csv
import io
import re
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime

from qgis.PyQt.QtCore import (
    Qt, QObject, QEvent, QFileSystemWatcher, QSettings, QTimer,
    QSize, QUrl, QPoint
)
from qgis.PyQt.QtWidgets import (
    QWidget, QLabel, QPushButton, QHBoxLayout,
    QVBoxLayout, QToolBar, QApplication, QSizePolicy,
    QFileDialog, QInputDialog, QMessageBox, QMenu, QDialog,
    QTextEdit, QLineEdit, QDialogButtonBox
)
from qgis.PyQt.QtGui import QFont, QIcon, QPixmap, QDesktopServices, QFontMetrics

from qgis.core import QgsMessageLog, Qgis

try:
    import openpyxl
except Exception:
    openpyxl = None


class WheelFilter(QObject):
    def __init__(self, parent_handler):
        super().__init__()
        self.parent_handler = parent_handler

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel:
            delta = event.angleDelta().y()
            if delta < 0:
                try:
                    self.parent_handler.next_entry()
                except Exception:
                    pass
            elif delta > 0:
                try:
                    self.parent_handler.prev_entry()
                except Exception:
                    pass
            return True
        return False


class Halo:
    def __init__(self, iface):
        self.iface = iface
        self.toolbar = None
        self.widget = None

        self.icon_btn = None
        self.link_btns = [None, None, None]
        self.num_btn = None
        self.date_label = None
        self.nav_widget = None
        self.btn_up = None
        self.btn_down = None
        self.msg_label = None
        self.right_pane = None
        self.unread_btn = None
        self.mark_all_btn = None
        self.mark_all_folder_label = None
        self.mark_all_tick_label = None
        self.halo_icon_label = None
        self.add_btn = None

        self.entries = []
        self.index = 0
        self.read_flags = []

        self.filepath = ''
        self.fs_watcher = QFileSystemWatcher()
        self.settings = QSettings()
        self._settings_key = 'Halo/filepath'
        self._index_key = 'Halo/index'
        self._read_map_key = 'Halo/read_map'
        self._signature_key = 'Halo/signature'
        self._external_link_keys = [
            'Halo/external_link_1',
            'Halo/external_link_2',
            'Halo/external_link_3'
        ]
        self._webapp_url_key = 'Halo/webapp_url'
        self._webapp_token_key = 'Halo/webapp_token'
        self._form_url_key = 'Halo/form_url'

        self.blink_timer = QTimer()
        self.blink_timer.setInterval(600)
        self.blink_timer.timeout.connect(self._on_blink_timeout)
        self._blink_state = False

        self.halo_blink_timer = QTimer()
        self.halo_blink_timer.setInterval(700)
        self.halo_blink_timer.timeout.connect(self._on_halo_blink)
        self._halo_blink_state = False

        self.auto_timer = QTimer()
        self.auto_timer.setInterval(60 * 1000)
        self.auto_timer.timeout.connect(self._on_auto_refresh)
        self._is_refreshing = False

        self._halo_colored = None
        self._months_pl = ["sty", "lut", "mar", "kwi", "maj", "cze", "lip", "sie", "wrz", "paź", "lis", "gru"]
        self._last_unread_pos = -1

    def _log(self, message: str, level=Qgis.Info):
        try:
            QgsMessageLog.logMessage(str(message), 'Halo', level)
        except Exception:
            pass

    def initGui(self):
        saved_fp = self.settings.value(self._settings_key, '')
        if saved_fp:
            self.filepath = saved_fp
        try:
            self.index = int(self.settings.value(self._index_key, 0))
        except Exception:
            self.index = 0

        self.toolbar = QToolBar("Halo")
        self.toolbar.setMovable(True)
        self.toolbar.setFloatable(True)
        self.toolbar.setObjectName("HaloToolbar")

        self.widget = QWidget()
        layout = QHBoxLayout()
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(6)

        screen = QApplication.primaryScreen()
        dpi_x = screen.logicalDotsPerInchX() if screen else 96.0
        cm_to_px = lambda cm: int(round((cm * dpi_x) / 2.54))

        first_w = cm_to_px(1.5)
        second_w = cm_to_px(12) - 100
        if second_w < cm_to_px(6):
            second_w = cm_to_px(6)
        third_w = cm_to_px(6)

        self.icon_btn = QPushButton()
        self.icon_btn.setToolTip("Wybierz plik z komunikatami lub wklej URL (prawy przycisk) / Ustaw WebApp")
        self.icon_btn.setFixedSize(40, 40)
        self.icon_btn.setIconSize(QSize(34, 34))
        self.icon_btn.setStyleSheet("padding:2px;")
        try:
            plugin_dir = os.path.dirname(__file__)
            icon_path = os.path.join(plugin_dir, 'icon.png')
            if os.path.exists(icon_path):
                self.icon_btn.setIcon(QIcon(icon_path))
                self._load_halo_pixmap(icon_path)
        except Exception:
            pass
        self.icon_btn.clicked.connect(self.choose_file)
        self.icon_btn.setContextMenuPolicy(Qt.CustomContextMenu)
        self.icon_btn.customContextMenuRequested.connect(self._icon_context_menu)

        link_btn_w = max(16, int(self.icon_btn.sizeHint().width() / 2))
        link_btn_h = self.icon_btn.sizeHint().height()
        colors = ['red', 'green', 'blue']
        for i, color in enumerate(colors):
            btn = QPushButton("♦")
            btn.setFixedSize(link_btn_w, link_btn_h)
            font_link = QFont()
            font_link.setPointSize(12)
            font_link.setBold(True)
            btn.setFont(font_link)
            btn.setStyleSheet(f"color: {color};")
            btn.clicked.connect(lambda _, idx=i: self._on_link_left_click(idx))
            btn.setContextMenuPolicy(Qt.CustomContextMenu)
            btn.customContextMenuRequested.connect(lambda pos, idx=i: self._on_link_right_click(idx, pos))
            btn.setToolTip("Lewy klik: otwórz. Prawy klik: ustaw.")
            self.link_btns[i] = btn

        self.nav_widget = QWidget()
        nav_layout = QVBoxLayout()
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(2)
        self.btn_up = QPushButton("▲")
        self.btn_up.setToolTip("Poprzedni komunikat")
        self.btn_down = QPushButton("▼")
        self.btn_down.setToolTip("Następny komunikat (jeśli nie ostatni, miga)")
        self.btn_up.setFixedWidth(18)
        self.btn_down.setFixedWidth(18)
        self.btn_up.setFixedHeight(18)
        self.btn_down.setFixedHeight(18)
        nav_layout.addWidget(self.btn_up, alignment=Qt.AlignHCenter | Qt.AlignTop)
        nav_layout.addWidget(self.btn_down, alignment=Qt.AlignHCenter | Qt.AlignTop)
        nav_layout.addStretch()
        self.nav_widget.setLayout(nav_layout)

        num_col_widget = QWidget()
        num_col_layout = QVBoxLayout()
        num_col_layout.setContentsMargins(0, 0, 0, 0)
        num_col_layout.setSpacing(2)

        self.num_btn = QPushButton("-")
        self.num_btn.setToolTip("Kliknij, aby oznaczyć jako przeczytane/nieprzeczytane")
        self.num_btn.setFlat(True)
        font_num = QFont()
        font_num.setBold(True)
        font_num.setPointSize(14)
        self.num_btn.setFont(font_num)
        self.num_btn.setObjectName("halo_num_btn")
        self.num_btn.setMinimumWidth(first_w)
        self.num_btn.setMaximumWidth(first_w)
        self.num_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
        self.num_btn.clicked.connect(self._on_num_clicked)

        self.date_label = QLabel("")
        font_date = QFont()
        font_date.setPointSize(9)
        self.date_label.setFont(font_date)
        self.date_label.setAlignment(Qt.AlignCenter)

        num_col_layout.addWidget(self.num_btn, 0, Qt.AlignTop)
        num_col_layout.addWidget(self.date_label, 0, Qt.AlignTop)
        num_col_layout.addStretch()
        num_col_widget.setLayout(num_col_layout)

        self.msg_label = QLabel("Brak danych")
        self.msg_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.msg_label.setWordWrap(True)
        font_msg = QFont()
        font_msg.setPointSize(10)
        self.msg_label.setFont(font_msg)
        self.msg_label.setMinimumWidth(second_w)
        self.msg_label.setMaximumWidth(second_w)
        self.msg_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)

        self.unread_btn = QPushButton("Nieprzeczytane: 0")
        self.unread_btn.setFlat(True)
        self.unread_btn.setFixedWidth(170)
        font_unread = QFont()
        font_unread.setBold(True)
        font_unread.setPointSize(10)
        self.unread_btn.setFont(font_unread)
        self.unread_btn.clicked.connect(self._on_unread_clicked)
        self.unread_btn.setToolTip("Kliknij: przejdź do kolejnego nieprzeczytanego komunikatu")

        self.right_pane = QWidget()
        rp_layout = QVBoxLayout()
        rp_layout.setContentsMargins(0, 0, 0, 0)
        rp_layout.setSpacing(4)

        top_row = QWidget()
        top_row_layout = QHBoxLayout()
        top_row_layout.setContentsMargins(0, 0, 0, 0)
        top_row_layout.setSpacing(8)

        self.mark_all_btn = QPushButton()
        self.mark_all_btn.setToolTip("Oznacz wszystkie jako przeczytane")
        self.mark_all_btn.setFlat(False)
        self.mark_all_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: 1px solid rgba(0,0,0,0);
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: rgba(200, 0, 0, 0.06);
            }
            QPushButton:pressed {
                background-color: rgba(200, 0, 0, 0.12);
            }
        """)
        mark_w = 34
        mark_h = 34
        self.mark_all_btn.setFixedSize(mark_w, mark_h)

        folder_lbl = QLabel("🗂️", self.mark_all_btn)
        folder_lbl.setAttribute(Qt.WA_TransparentForMouseEvents)
        folder_lbl.setAlignment(Qt.AlignCenter)
        font_mark_folder = QFont()
        font_mark_folder.setPointSize(14)
        folder_lbl.setFont(font_mark_folder)
        folder_lbl.setGeometry(0, 0, mark_w, mark_h)

        tick_lbl = QLabel("✔", self.mark_all_btn)
        tick_lbl.setAttribute(Qt.WA_TransparentForMouseEvents)
        tick_lbl_font = QFont()
        tick_lbl_font.setPointSize(10)
        tick_lbl_font.setBold(True)
        tick_lbl.setFont(tick_lbl_font)
        tick_lbl.setStyleSheet("color: red; background: rgba(0,0,0,0);")
        tick_x = int(mark_w * 0.58)
        tick_y = int(mark_h * 0.12)
        tick_lbl.move(tick_x, tick_y)
        tick_lbl.resize(12, 12)

        self.mark_all_folder_label = folder_lbl
        self.mark_all_tick_label = tick_lbl
        self.mark_all_btn.clicked.connect(self._on_mark_all_clicked)

        self.add_btn = QPushButton("🚀")
        self.add_btn.setToolTip("Lewy klik: otwórz formularz (jeśli ustawiono) lub Dodaj komunikat. Prawy klik: ustaw link formularza.")
        font_plus = QFont()
        font_plus.setPointSize(18)
        font_plus.setBold(True)
        self.add_btn.setFont(font_plus)
        self.add_btn.setFixedSize(36, 36)
        self.add_btn.setStyleSheet("""
            QPushButton { background: transparent; border-radius: 6px; }
            QPushButton:hover { background-color: rgba(0, 122, 204, 0.08); }
        """)
        self.add_btn.clicked.connect(self._on_add_btn_left)
        self.add_btn.setContextMenuPolicy(Qt.CustomContextMenu)
        self.add_btn.customContextMenuRequested.connect(self._add_btn_context_menu)

        self.halo_icon_label = QLabel()
        self.halo_icon_label.setFixedSize(32, 32)
        self.halo_icon_label.setScaledContents(True)

        top_row_layout.addWidget(self.mark_all_btn, alignment=Qt.AlignVCenter)
        top_row_layout.addWidget(self.add_btn, alignment=Qt.AlignVCenter)
        top_row_layout.addWidget(self.halo_icon_label, alignment=Qt.AlignVCenter)
        top_row_layout.addWidget(self.unread_btn, alignment=Qt.AlignVCenter)
        top_row_layout.addStretch()
        top_row.setLayout(top_row_layout)

        rp_layout.addStretch()
        rp_layout.addWidget(top_row, alignment=Qt.AlignHCenter)
        rp_layout.addStretch()
        self.right_pane.setLayout(rp_layout)

        # tylko ta zmiana: +100 px szerokości prawego panelu
        right_pane_w = third_w + 100
        self.right_pane.setMinimumWidth(right_pane_w)
        self.right_pane.setMaximumWidth(right_pane_w)

        layout.addWidget(self.icon_btn, 0, Qt.AlignTop)
        for b in self.link_btns:
            layout.addWidget(b, 0, Qt.AlignTop)
        layout.addWidget(self.nav_widget, 0, Qt.AlignTop)
        layout.addWidget(num_col_widget, 0, Qt.AlignTop)
        layout.addWidget(self.msg_label, 0, Qt.AlignTop)
        layout.addWidget(self.right_pane, 0, Qt.AlignTop)

        self.widget.setLayout(layout)
        self.toolbar.addWidget(self.widget)
        self.iface.mainWindow().addToolBar(self.toolbar)

        self.btn_up.clicked.connect(self.prev_entry)
        self.btn_down.clicked.connect(self.next_entry)

        self.wheel_filter = WheelFilter(self)
        self.num_btn.installEventFilter(self.wheel_filter)
        self.msg_label.installEventFilter(self.wheel_filter)

        try:
            if self.filepath and os.path.exists(self.filepath):
                self.fs_watcher.addPath(self.filepath)
            self.fs_watcher.fileChanged.connect(self._on_file_changed)
        except Exception:
            pass

        self._load_saved_links()

        try:
            form_link = str(self.settings.value(self._form_url_key, '') or '')
            if form_link:
                self.add_btn.setToolTip(f"Lewy klik: otwórz formularz: {form_link}\nPrawy klik: ustaw link formularza.")
        except Exception:
            pass

        try:
            weburl = self.settings.value(self._webapp_url_key, '')
            if weburl:
                self.auto_timer.start()
        except Exception:
            pass

        self.reload_entries()

    def choose_file(self):
        try:
            pos = self.icon_btn.rect().bottomLeft()
            self._icon_context_menu(pos)
        except Exception:
            try:
                start_dir = os.path.dirname(self.filepath) if self.filepath and os.path.exists(self.filepath) else os.path.expanduser('~')
                fn, _ = QFileDialog.getOpenFileName(
                    self.iface.mainWindow(),
                    'Wybierz plik z komunikatami',
                    start_dir,
                    'All files (*);;Text files (*.txt);;CSV files (*.csv);;Excel files (*.xlsx *.xls)'
                )
                if fn:
                    self._set_new_source(fn)
            except Exception:
                pass

    def _icon_context_menu(self, pos):
        menu = QMenu(self.icon_btn)
        act_file = menu.addAction("Wybierz plik lokalny...")
        act_url = menu.addAction("Wprowadź URL źródła (txt/csv/xlsx/gsheet)...")
        act_set_webapp = menu.addAction("Ustaw WebApp URL / Token")
        act_clear = menu.addAction("Wyczyść źródło")
        chosen = menu.exec_(self.icon_btn.mapToGlobal(pos))
        if chosen == act_file:
            start_dir = os.path.dirname(self.filepath) if self.filepath and os.path.exists(self.filepath) else os.path.expanduser('~')
            fn, _ = QFileDialog.getOpenFileName(
                self.iface.mainWindow(),
                'Wybierz plik z komunikatami',
                start_dir,
                'All files (*);;Text files (*.txt);;CSV files (*.csv);;Excel files (*.xlsx *.xls)'
            )
            if fn:
                self._set_new_source(fn)
        elif chosen == act_url:
            current = self.settings.value(self._settings_key, self.filepath)
            text, ok = QInputDialog.getText(
                self.iface.mainWindow(),
                "Wprowadź URL źródła",
                "Wklej URL (txt/csv/xlsx/google sheet):",
                text=current
            )
            if ok:
                url = text.strip()
                if url:
                    self._set_new_source(url)
        elif chosen == act_set_webapp:
            self._webapp_settings_dialog()
        elif chosen == act_clear:
            self._set_new_source('')

    def _webapp_settings_dialog(self):
        dlg = QDialog(self.iface.mainWindow())
        dlg.setWindowTitle("Ustawienia WebApp")
        v = QVBoxLayout()
        url_edit = QLineEdit()
        url_edit.setPlaceholderText("Wklej URL WebApp (np. https://script.google.com/.../exec)")
        url_edit.setText(str(self.settings.value(self._webapp_url_key, '')))
        v.addWidget(QLabel("WebApp URL:"))
        v.addWidget(url_edit)
        token_edit = QLineEdit()
        token_edit.setPlaceholderText("SECRET TOKEN (zgodny z WebApp)")
        token_edit.setText(str(self.settings.value(self._webapp_token_key, '')))
        v.addWidget(QLabel("Token:"))
        v.addWidget(token_edit)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        v.addWidget(bb)
        dlg.setLayout(v)

        def accept():
            u = url_edit.text().strip()
            t = token_edit.text().strip()
            self.settings.setValue(self._webapp_url_key, u)
            self.settings.setValue(self._webapp_token_key, t)
            try:
                if u:
                    self.auto_timer.start()
                else:
                    if self.auto_timer.isActive():
                        self.auto_timer.stop()
            except Exception:
                pass
            dlg.accept()

        bb.accepted.connect(accept)
        bb.rejected.connect(lambda: dlg.reject())
        dlg.exec_()

    def _set_new_source(self, path_or_url):
        try:
            old = self.filepath
            try:
                if old and os.path.exists(old) and old in self.fs_watcher.files():
                    self.fs_watcher.removePath(old)
            except Exception:
                pass

            self.filepath = path_or_url

            try:
                if self.filepath and os.path.exists(self.filepath):
                    self.fs_watcher.addPath(self.filepath)
            except Exception:
                pass

            self.settings.setValue(self._settings_key, self.filepath)
            self.index = 0
            self._save_index()

            try:
                if self.filepath:
                    self.auto_timer.start()
                else:
                    if self.auto_timer.isActive():
                        self.auto_timer.stop()
            except Exception:
                pass

            self.reload_entries()
            self._log(f"Nowe źródło ustawione: {path_or_url}")
        except Exception:
            pass

    def unload(self):
        try:
            if self.toolbar:
                self.iface.mainWindow().removeToolBar(self.toolbar)
        except Exception:
            pass
        try:
            if self.blink_timer.isActive():
                self.blink_timer.stop()
        except Exception:
            pass
        try:
            if self.halo_blink_timer.isActive():
                self.halo_blink_timer.stop()
        except Exception:
            pass
        try:
            if self.auto_timer.isActive():
                self.auto_timer.stop()
        except Exception:
            pass
        try:
            self.fs_watcher.removePaths(self.fs_watcher.files())
        except Exception:
            pass

    def _add_btn_context_menu(self, pos):
        menu = QMenu(self.add_btn)
        act_set_form = menu.addAction("Ustaw link formularza respondenta...")
        act_clear = menu.addAction("Wyczyść link formularza")
        chosen = menu.exec_(self.add_btn.mapToGlobal(pos))
        if chosen == act_set_form:
            current = self.settings.value(self._form_url_key, '')
            text, ok = QInputDialog.getText(
                self.iface.mainWindow(),
                "Wprowadź link formularza",
                "Wklej URL formularza Google Forms (viewform):",
                text=str(current)
            )
            if ok:
                u = text.strip()
                self.settings.setValue(self._form_url_key, u)
                if u:
                    self.add_btn.setToolTip(f"Lewy klik: otwórz formularz: {u}\nPrawy klik: ustaw link formularza.")
                else:
                    self.add_btn.setToolTip(
                        "Lewy klik: otwórz formularz (jeśli ustawiono) lub Dodaj komunikat. Prawy klik: ustaw link formularza."
                    )
        elif chosen == act_clear:
            self.settings.setValue(self._form_url_key, '')
            self.add_btn.setToolTip(
                "Lewy klik: otwórz formularz (jeśli ustawiono) lub Dodaj komunikat. Prawy klik: ustaw link formularza."
            )

    def _on_add_btn_left(self):
        try:
            form_url = str(self.settings.value(self._form_url_key, '') or '').strip()
            if form_url:
                QDesktopServices.openUrl(QUrl(form_url))
                return
        except Exception:
            pass
        self._on_add_clicked()

    def _load_saved_links(self):
        try:
            for i, key in enumerate(self._external_link_keys):
                link = self.settings.value(key, '')
                if i == 0 and not link:
                    link = "https://nowiny24.pl/"
                    self.settings.setValue(key, link)
                if link:
                    self.link_btns[i].setToolTip(f"Lewy klik: otwórz\nPrawy klik: ustaw\n\n{link}")
                else:
                    self.link_btns[i].setToolTip("Lewy klik: otwórz. Prawy klik: ustaw.")
        except Exception:
            pass

    def _save_link_for_index(self, idx, link_text):
        try:
            key = self._external_link_keys[idx]
            self.settings.setValue(key, link_text)
            if link_text:
                self.link_btns[idx].setToolTip(f"Lewy klik: otwórz\nPrawy klik: ustaw\n\n{link_text}")
            else:
                self.link_btns[idx].setToolTip("Lewy klik: otwórz. Prawy klik: ustaw.")
        except Exception:
            pass

    def _on_link_right_click(self, idx: int, pos: QPoint):
        try:
            btn = self.link_btns[idx]
            menu = QMenu(btn)
            act_type = menu.addAction("Wpisz link/ścieżkę...")
            act_file = menu.addAction("Wybierz plik...")
            act_folder = menu.addAction("Wybierz folder...")
            act_clear = menu.addAction("Wyczyść przypisanie")
            chosen = menu.exec_(btn.mapToGlobal(pos))
            if chosen == act_type:
                current = self.settings.value(self._external_link_keys[idx], '')
                text, ok = QInputDialog.getText(
                    self.iface.mainWindow(),
                    "Wpisz link/ścieżkę",
                    "Wprowadź link lub ścieżkę:",
                    text=current
                )
                if ok:
                    self._save_link_for_index(idx, text.strip())
            elif chosen == act_file:
                start_dir = os.path.dirname(self.settings.value(self._external_link_keys[idx], '')) or os.path.expanduser('~')
                fn, _ = QFileDialog.getOpenFileName(self.iface.mainWindow(), "Wybierz plik", start_dir, "Wszystkie pliki (*)")
                if fn:
                    self._save_link_for_index(idx, fn)
            elif chosen == act_folder:
                start_dir = os.path.dirname(self.settings.value(self._external_link_keys[idx], '')) or os.path.expanduser('~')
                d = QFileDialog.getExistingDirectory(self.iface.mainWindow(), "Wybierz folder", start_dir)
                if d:
                    self._save_link_for_index(idx, d)
            elif chosen == act_clear:
                self._save_link_for_index(idx, '')
        except Exception:
            pass

    def _on_link_left_click(self, idx: int):
        try:
            key = self._external_link_keys[idx]
            stored = self.settings.value(key, '')
            if not stored:
                QMessageBox.information(
                    self.iface.mainWindow(),
                    "Brak przypisania",
                    "Nie ustawiono przypisania dla tego przycisku. Ustaw prawym przyciskiem myszy."
                )
                return

            candidate = stored.strip()
            candidate_expanded = os.path.expanduser(candidate)
            candidate_norm = os.path.normpath(candidate_expanded)

            if os.path.exists(candidate_norm):
                QDesktopServices.openUrl(QUrl.fromLocalFile(candidate_norm))
                return

            q = QUrl(candidate)
            if q.isValid() and q.scheme():
                QDesktopServices.openUrl(q)
                return

            if "://" not in candidate:
                candidate_http = "http://" + candidate
            else:
                candidate_http = candidate

            q2 = QUrl(candidate_http)
            if q2.isValid():
                QDesktopServices.openUrl(q2)
                return

            QMessageBox.warning(self.iface.mainWindow(), "Nie można otworzyć", f"Nie udało się otworzyć przypisania:\n{stored}")
        except Exception:
            QMessageBox.warning(self.iface.mainWindow(), "Błąd", "Nie udało się otworzyć przypisania.")

    def _looks_like_spreadsheet(self, path_or_url):
        lower = (path_or_url or "").lower()
        if lower.startswith("http://") or lower.startswith("https://") or lower.endswith(".csv"):
            return True
        for ext in ('.xlsx', '.xls', '.csv'):
            if lower.endswith(ext):
                return True
        return False

    def _load_from_spreadsheet(self, path_or_url):
        is_google_sheet = (
            (path_or_url or '').lower().startswith(('http://', 'https://'))
            and 'docs.google.com/spreadsheets' in (path_or_url or '').lower()
        )
        form_configured = bool(str(self.settings.value(self._form_url_key, '') or '').strip())
        try:
            if (path_or_url or "").lower().startswith(('http://', 'https://')):
                result = self._fetch_csv_from_url(path_or_url)
                if result is None:
                    return []
                kind, data = result

                if is_google_sheet and form_configured:
                    if kind == 'csv':
                        text = data.decode('utf-8-sig', errors='replace')
                        rows = list(csv.reader(io.StringIO(text)))
                        return self._entries_from_forms_responses(rows)
                    elif kind == 'xlsx':
                        if openpyxl is None:
                            self.msg_label.setText("Plik XLSX pobrany, ale brak biblioteki openpyxl")
                            return []
                        try:
                            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                            ws = wb.active
                            rows = list(ws.iter_rows(values_only=True))
                            return self._entries_from_forms_responses(rows)
                        except Exception as e:
                            self.msg_label.setText("Błąd parsowania XLSX z URL")
                            self._log(f"Błąd parsowania XLSX z URL: {e}", Qgis.Critical)
                            return []
                    else:
                        self.msg_label.setText("Nieznany typ pliku z URL")
                        self._log("Nieznany typ pliku z URL", Qgis.Warning)
                        return []

                if kind == 'csv':
                    text = data.decode('utf-8-sig', errors='replace')
                    return self._entries_from_dictreader(csv.DictReader(io.StringIO(text)))
                elif kind == 'xlsx':
                    if openpyxl is None:
                        self.msg_label.setText("Plik XLSX pobrany, ale brak biblioteki openpyxl")
                        return []
                    try:
                        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                        ws = wb.active
                        rows = list(ws.iter_rows(values_only=True))
                        return self._entries_from_rows(rows)
                    except Exception as e:
                        self.msg_label.setText("Błąd parsowania XLSX z URL")
                        self._log(f"Błąd parsowania XLSX z URL: {e}", Qgis.Critical)
                        return []
                else:
                    self.msg_label.setText("Nieznany typ pliku z URL")
                    self._log("Nieznany typ pliku z URL", Qgis.Warning)
                    return []

            lower = path_or_url.lower()
            if lower.endswith('.csv'):
                try:
                    with open(path_or_url, 'r', encoding='utf-8', newline='') as fh:
                        return self._entries_from_dictreader(csv.DictReader(fh))
                except Exception:
                    try:
                        with open(path_or_url, 'r', encoding='cp1250', newline='') as fh:
                            return self._entries_from_dictreader(csv.DictReader(fh))
                    except Exception as e:
                        self.msg_label.setText("Błąd czytania CSV")
                        self._log(f"Błąd czytania CSV: {e}", Qgis.Warning)
                        return []

            if lower.endswith(('.xlsx', '.xls')):
                if openpyxl is None:
                    self.num_btn.setText('#')
                    self.msg_label.setText("Brak obsługi plików Excel: brakuje biblioteki openpyxl")
                    self._log("Brak biblioteki openpyxl (lokalne XLSX)", Qgis.Warning)
                    return []
                try:
                    wb = openpyxl.load_workbook(path_or_url, read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if form_configured and 'docs.google.com/spreadsheets' in str(path_or_url).lower():
                        return self._entries_from_forms_responses(rows)
                    return self._entries_from_rows(rows)
                except Exception as e:
                    self.msg_label.setText("Błąd czytania pliku Excel")
                    self._log(f"Błąd czytania pliku Excel: {e}", Qgis.Critical)
                    return []

            return []
        except Exception as e:
            self.msg_label.setText("Błąd pobierania/parsingu arkusza")
            self._log(f"Błąd _load_from_spreadsheet: {e}", Qgis.Critical)
            return []

    def _entries_from_forms_responses(self, rows):
        entries = []
        if not rows or len(rows) < 2:
            return entries
        for idx, r in enumerate(rows):
            if idx == 0:
                continue
            try:
                timestamp = ''
                text = ''
                signature = ''

                if isinstance(r, (list, tuple)):
                    if len(r) > 0 and r[0] is not None:
                        timestamp = str(r[0]).strip()
                    if len(r) > 1 and r[1] is not None:
                        text = str(r[1])
                    if len(r) > 2 and r[2] is not None:
                        signature = str(r[2]).strip()
                else:
                    text = str(r) if r is not None else ''

                sheet_row_number = idx + 1
                try:
                    num_val = int(sheet_row_number) - 1
                    if num_val < 1:
                        num_val = 1
                    num = str(num_val)
                except Exception:
                    num = str(max(1, sheet_row_number - 1))

                dt = self._try_parse_any_datetime(timestamp)
                msg_html = f"{str(text)} <i>~{signature}</i>" if signature else f"{str(text)}"
                entries.append((num, dt, timestamp, msg_html))
            except Exception:
                continue
        return entries

    def _entries_from_rows(self, rows):
        if not rows:
            return []
        header_row_idx = None
        headers = None

        for idx, row in enumerate(rows):
            if not row:
                continue
            normalized = [(str(c).strip() if c is not None else '').lower() for c in row]
            text_join = ' '.join(normalized)
            if any(tok in text_join for tok in ('nr', 'no', 'number', 'lp', 'id', 'time', 'text', 'message', 'treść', 'tresc', 'date', 'czas')):
                header_row_idx = idx
                headers = normalized
                break

        if headers is None:
            headers = [(str(c).strip() if c is not None else '').lower() for c in rows[0]]
            header_row_idx = 0

        col_map = {}
        for i, h in enumerate(headers):
            if h in ('nr', 'no', 'number', 'lp', 'id', 'idnr'):
                col_map['nr'] = i
            elif h in ('time', 'timestamp', 'date', 'czas'):
                col_map['time'] = i
            elif h in ('text', 'message', 'treść', 'tresc'):
                col_map['text'] = i

        if 'nr' not in col_map:
            for i, h in enumerate(headers):
                if any(p in h for p in ('nr', 'no', 'num', 'lp', 'id')):
                    col_map['nr'] = i
                    break
        if 'time' not in col_map:
            for i, h in enumerate(headers):
                if any(p in h for p in ('time', 'date', 'czas', 'timestamp')):
                    col_map['time'] = i
                    break
        if 'text' not in col_map:
            for i, h in enumerate(headers):
                if any(p in h for p in ('text', 'message', 'tre', 'tresc')):
                    col_map['text'] = i
                    break

        entries = []
        data_rows = rows[header_row_idx + 1:]
        for ridx, r in enumerate(data_rows):
            if r is None:
                continue
            if all((c is None or str(c).strip() == '') for c in r):
                continue

            num = ''
            raw_time = ''
            dt = None

            if 'nr' in col_map and col_map['nr'] < len(r):
                val = r[col_map['nr']]
                if val is not None:
                    num = str(val).strip()

            if not num:
                for c in r:
                    if c is None:
                        continue
                    s = str(c).strip()
                    if re.match(r'^\d+$', s):
                        num = s
                        break
            if not num:
                num = str(ridx + 1)

            if 'time' in col_map and col_map['time'] < len(r):
                rt = r[col_map['time']]
                if isinstance(rt, datetime):
                    dt = rt
                    raw_time = dt.strftime("%d.%m.%Y %H:%M")
                elif rt is not None:
                    raw_time = str(rt).strip()
                    dt = self._try_parse_any_datetime(raw_time)

            msg = ''
            if 'text' in col_map and col_map['text'] < len(r):
                tv = r[col_map['text']]
                msg = '' if tv is None else str(tv)
            else:
                for c in reversed(r):
                    if c is not None and str(c).strip():
                        msg = str(c)
                        break

            entries.append((num, dt, raw_time, msg))
        return entries

    def _entries_from_dictreader(self, reader):
        fieldnames = [fn.strip() for fn in (reader.fieldnames or [])]
        norm = [fn.lower() for fn in fieldnames]
        idx_nr = None
        idx_time = None
        idx_text = None

        for i, name in enumerate(norm):
            if name in ('nr', 'no', 'number', 'lp', 'id'):
                idx_nr = fieldnames[i]
            elif name in ('time', 'timestamp', 'date', 'czas'):
                idx_time = fieldnames[i]
            elif name in ('text', 'message', 'treść', 'tresc'):
                idx_text = fieldnames[i]

        if idx_nr is None:
            for i, name in enumerate(norm):
                if any(p in name for p in ('nr', 'no', 'num', 'lp', 'id')):
                    idx_nr = fieldnames[i]
                    break
        if idx_time is None:
            for i, name in enumerate(norm):
                if any(p in name for p in ('time', 'date', 'czas', 'timestamp', 'data')):
                    idx_time = fieldnames[i]
                    break
        if idx_text is None:
            for i, name in enumerate(norm):
                if any(p in name for p in ('text', 'message', 'tre', 'tresc', 'komunikat')):
                    idx_text = fieldnames[i]
                    break

        entries = []
        row_index = 0
        for row in reader:
            try:
                row_index += 1
                num = ''
                raw_time_str = ''
                dt = None
                msg = ''

                if idx_nr and idx_nr in row:
                    num = str(row[idx_nr]).strip()
                if not num:
                    for key, val in row.items():
                        if val is None:
                            continue
                        s = str(val).strip()
                        if re.match(r'^\d+$', s):
                            num = s
                            break
                if not num:
                    num = str(row_index)

                if idx_time and idx_time in row:
                    raw_time_str = str(row[idx_time]).strip()
                    dt = self._try_parse_any_datetime(raw_time_str)
                else:
                    raw_time_str = ''
                    dt = None

                if idx_text and idx_text in row:
                    v = row[idx_text]
                    msg = '' if v is None else str(v)
                else:
                    if fieldnames:
                        last = fieldnames[-1]
                        v = row.get(last, '')
                        msg = '' if v is None else str(v)
                    else:
                        msg = ''

                entries.append((num, dt, raw_time_str, msg))
            except Exception:
                continue
        return entries

    def _try_parse_any_datetime(self, text):
        if text is None:
            return None
        txt = str(text).strip()
        if not txt:
            return None

        fmts = [
            "%H:%M %d.%m.%Y",
            "%d.%m.%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y",
            "%H:%M",
            "%H:%M:%S",
            "%d.%m.%Y",
            "%Y.%m.%d %H:%M",
            "%d-%m-%Y %H:%M"
        ]
        for f in fmts:
            try:
                return datetime.strptime(txt, f)
            except Exception:
                continue

        try:
            return datetime.fromisoformat(txt.replace('Z', ''))
        except Exception:
            return None

    def _parse_file(self):
        entries = []
        try:
            if not os.path.exists(self.filepath):
                return entries
            with open(self.filepath, "r", encoding="utf-8") as f:
                lines = [ln.rstrip('\n') for ln in f]

            block = []
            for line in lines:
                if line.strip() == '***':
                    if block:
                        first_line = None
                        first_idx = None
                        for idx, bl in enumerate(block):
                            if bl.strip():
                                first_line = bl
                                first_idx = idx
                                break
                        if first_line is not None:
                            parts = first_line.split(';', 2)
                            if len(parts) >= 3:
                                num = parts[0].strip()
                                raw_date = parts[1].strip()
                                first_msg_part = parts[2].rstrip()
                                remaining = block[first_idx + 1:] if first_idx is not None and first_idx + 1 < len(block) else []
                                msg = '\n'.join([first_msg_part] + remaining).strip()
                                dt = self._try_parse_any_datetime(raw_date)
                                entries.append((num, dt, raw_date, msg))
                        block = []
                else:
                    block.append(line)

            if block:
                first_line = None
                first_idx = None
                for idx, bl in enumerate(block):
                    if bl.strip():
                        first_line = bl
                        first_idx = idx
                        break
                if first_line is not None:
                    parts = first_line.split(';', 2)
                    if len(parts) >= 3:
                        num = parts[0].strip()
                        raw_date = parts[1].strip()
                        first_msg_part = parts[2].rstrip()
                        remaining = block[first_idx + 1:] if first_idx is not None and first_idx + 1 < len(block) else []
                        msg = '\n'.join([first_msg_part] + remaining).strip()
                        dt = self._try_parse_any_datetime(raw_date)
                        entries.append((num, dt, raw_date, msg))
        except Exception:
            pass
        return entries

    def _fetch_csv_from_url(self, url):
        debug = []
        try:
            u0 = (url or '').strip()
            debug.append(f"Źródłowy URL: {u0}")
            if not u0:
                self._log("Pusty URL źródła.", Qgis.Warning)
                return None

            candidates = []
            u = u0

            if 'docs.google.com/spreadsheets' in u:
                m = re.search(r'/d/([a-zA-Z0-9-_]+)', u)
                sid = m.group(1) if m else None
                gid = None
                mg = re.search(r'[?&]gid=(\d+)', u)
                if mg:
                    gid = mg.group(1)
                else:
                    mfrag = re.search(r'#gid=(\d+)', u)
                    if mfrag:
                        gid = mfrag.group(1)
                if not gid:
                    gid = '0'
                if sid:
                    candidates.append(f'https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv&gid={gid}')
                    candidates.append(f'https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid}')
                    candidates.append(f'https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx&gid={gid}')
                    candidates.append(f'https://docs.google.com/feeds/download/spreadsheets/Export?key={sid}&exportFormat=csv&gid={gid}')
                    debug.append("Dodano candidate Google (gviz/export/feeds).")

            if 'drive.google.com/file' in u or 'drive.google.com/open' in u or 'drive.google.com/uc' in u:
                m = re.search(r'/file/d/([a-zA-Z0-9-_]+)', u)
                if not m:
                    m = re.search(r'id=([a-zA-Z0-9-_]+)', u)
                if m:
                    fid = m.group(1)
                    candidates.append(f'https://docs.google.com/uc?export=download&id={fid}')
                    debug.append("Dodano candidate Google Drive uc?export=download.")

            if 'dropbox.com' in u:
                if 'dl=0' in u:
                    candidates.append(u.replace('dl=0', 'dl=1'))
                elif '?' not in u:
                    candidates.append(u + '?dl=1')
                else:
                    candidates.append(u)
                debug.append("Dodano candidate Dropbox.")

            if '1drv.ms' in u or 'onedrive.live.com' in u:
                if '?' in u:
                    candidates.append(u + '&download=1')
                else:
                    candidates.append(u + '?download=1')
                candidates.append(u)
                debug.append("Dodano candidate OneDrive/1drv (z download=1).")

            if re.search(r'\.csv($|\?)', u, re.I) or re.search(r'\.xlsx($|\?)', u, re.I) or re.search(r'\.xls($|\?)', u, re.I):
                candidates.insert(0, u)

            seen = set()
            final_candidates = []
            for c in candidates + [u]:
                if c and c not in seen:
                    final_candidates.append(c)
                    seen.add(c)

            debug.append(f"Candidates: {final_candidates}")

            for candidate in final_candidates:
                debug.append(f"Próbuję pobrać: {candidate}")
                try:
                    req = urllib.request.Request(candidate, headers={'User-Agent': 'Mozilla/5.0', 'Accept': '*/*'})
                    with urllib.request.urlopen(req, timeout=25) as resp:
                        code = None
                        try:
                            code = resp.getcode()
                        except Exception:
                            code = None

                        info = resp.info()
                        content_type = ''
                        try:
                            content_type = info.get_content_type()
                        except Exception:
                            content_type = info.get('Content-Type', '') or ''

                        data = resp.read()
                        length = len(data)
                        debug.append(f"HTTP {code}  Content-Type: {content_type}  Rozmiar: {length} bytes")

                        if code in (400, 401, 403):
                            if 'docs.google.com' in candidate:
                                self._log(f"Google zwrócił HTTP {code}. Upewnij się, że arkusz jest publicznie udostępniony (Share/Publish) lub użyj WebApp.", Qgis.Warning)
                            elif '1drv.ms' in candidate or 'onedrive.live.com' in candidate:
                                self._log(f"OneDrive zwrócił HTTP {code}. Upewnij się, że link jest publiczny (direct download).", Qgis.Warning)
                            debug.append(f"Otrzymano HTTP {code}, przechodzę dalej.")
                            continue

                        if data[:2] == b'PK':
                            self._log(f"Pobrano binarny plik (ZIP/XLSX) z {candidate}")
                            return ('xlsx', data)

                        if 'spreadsheetml' in content_type or 'vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                            self._log(f"Content-Type wskazuje XLSX: {content_type}")
                            return ('xlsx', data)

                        if 'text/csv' in content_type or 'text/plain' in content_type or 'application/csv' in content_type:
                            self._log(f"Pobrano CSV/tekst (Content-Type: {content_type}), długość {length} bajtów")
                            return ('csv', data)

                        head = data[:4096].lower()
                        if b'<html' in head or b'<!doctype' in head or b'<script' in head:
                            s = data.decode('utf-8', errors='replace')
                            debug.append("Otrzymano HTML — analizuję w poszukiwaniu odnośników do pliku/eksportu...")
                            m_export = re.search(r'href=["\']([^"\']*export[^"\']*format=(?:csv|xlsx)[^"\']*)["\']', s, re.I)
                            if m_export:
                                found = m_export.group(1)
                                if found.startswith('/'):
                                    found = 'https://docs.google.com' + found
                                debug.append(f"Znaleziono link eksportu w HTML: {found} — próbuję pobrać")
                                try:
                                    req2 = urllib.request.Request(found, headers={'User-Agent': 'Mozilla/5.0', 'Accept': '*/*'})
                                    with urllib.request.urlopen(req2, timeout=20) as r2:
                                        d2 = r2.read()
                                        if d2[:2] == b'PK':
                                            self._log("Export link zwrócił XLSX.")
                                            return ('xlsx', d2)
                                        self._log("Export link zwrócił CSV/tekst.")
                                        return ('csv', d2)
                                except Exception as e2:
                                    debug.append(f"Błąd pobierania znalezionego linku export: {e2}")

                            m_file = re.search(r'href=["\']([^"\']*\.(?:csv|xlsx|xls))["\']', s, re.I)
                            if m_file:
                                found = m_file.group(1)
                                if found.startswith('/'):
                                    parsed = urllib.parse.urlparse(candidate)
                                    base = f"{parsed.scheme}://{parsed.netloc}"
                                    found = base + found
                                debug.append(f"Znaleziono w HTML bezpośredni href: {found} — próbuję pobrać")
                                try:
                                    req2 = urllib.request.Request(found, headers={'User-Agent': 'Mozilla/5.0', 'Accept': '*/*'})
                                    with urllib.request.urlopen(req2, timeout=20) as r2:
                                        d2 = r2.read()
                                        if d2[:2] == b'PK':
                                            self._log("Pobrano XLSX z linku znalezionego w HTML.")
                                            return ('xlsx', d2)
                                        self._log("Pobrano CSV z linku znalezionego w HTML.")
                                        return ('csv', d2)
                                except Exception as e2:
                                    debug.append(f"Błąd pobierania linku do pliku w HTML: {e2}")

                            m_js = re.search(
                                r'(?:location\.replace|window\.location\.href|window\.location)\s*\(\s*[\'"]([^\'"]+)[\'"]\s*\)',
                                s,
                                re.I
                            )
                            if not m_js:
                                m_js = re.search(r'window\.location\s*=\s*[\'"]([^\'"]+)[\'"]', s, re.I)
                            if m_js:
                                found = m_js.group(1)
                                debug.append(f"Znaleziono JS-redirect: {found} — próbuję pobrać")
                                try:
                                    req2 = urllib.request.Request(found, headers={'User-Agent': 'Mozilla/5.0', 'Accept': '*/*'})
                                    with urllib.request.urlopen(req2, timeout=20) as r2:
                                        d2 = r2.read()
                                        if d2[:2] == b'PK':
                                            self._log("JS redirect zwrócił XLSX.")
                                            return ('xlsx', d2)
                                        s2 = d2.decode('utf-8', errors='replace')
                                        if ',' in s2 or '\n' in s2:
                                            self._log("JS redirect zwrócił CSV.")
                                            return ('csv', d2)
                                except Exception as e2:
                                    debug.append(f"Błąd podążania za JS-redirect: {e2}")

                            debug.append("HTML nie zawiera rozpoznawalnego linku eksportu/pliku (lub wymaga auth).")
                            continue

                        try:
                            s = data.decode('utf-8', errors='replace')
                            preview = s[:400].replace('\n', '\\n')
                            debug.append("Podgląd tekstowy: " + preview)
                            if ',' in s or '\n' in s:
                                self._log("Pobrane dane wyglądają jak CSV (znaleziono przecinek/nowy wiersz).")
                                return ('csv', data)
                        except Exception:
                            debug.append("Dekodowanie UTF-8 nie powiodło się; traktuję dane jako binarne.")
                        self._log("Nie rozpoznano typu; zwracam CSV jako fallback.")
                        return ('csv', data)

                except Exception as e:
                    debug.append(f"Błąd pobierania candidate {candidate}: {e}")
                    continue

            self._log("Nie udało się pobrać pliku z żadnego z kandydatów. Sprawdź uprawnienia/URL.", Qgis.Warning)
            if 'docs.google.com' in u0:
                self._log(
                    "Google zwrócił stronę zamiast pliku. Upewnij się, że arkusz jest udostępniony lub opublikowany (Publish to web) lub użyj WebApp.",
                    Qgis.Warning
                )
            if '1drv.ms' in u0 or 'onedrive.live.com' in u0:
                self._log("OneDrive może wymagać bezpośredniego linku do pobrania (direct).", Qgis.Warning)

            for d in debug:
                QgsMessageLog.logMessage(d, 'Halo', Qgis.Info)

            return None
        except Exception as ex:
            self._log(f"Wyjątek podczas pobierania URL: {ex}", Qgis.Critical)
            return None

    def _load_read_map(self):
        try:
            raw = self.settings.value(self._read_map_key, '')
            if not raw:
                return {}
            if isinstance(raw, str):
                return json.loads(raw)
            return dict(raw)
        except Exception:
            return {}

    def _save_read_map(self):
        try:
            mapping = {}
            for idx, e in enumerate(self.entries):
                num = e[0]
                flag = 1 if idx < len(self.read_flags) and self.read_flags[idx] else 0
                mapping[str(num)] = flag
            self.settings.setValue(self._read_map_key, json.dumps(mapping))
        except Exception:
            pass

    def _save_index(self):
        try:
            self.settings.setValue(self._index_key, int(self.index))
        except Exception:
            pass

    def _on_file_changed(self, path):
        self.reload_entries()

    def reload_entries(self):
        self._last_unread_pos = -1

        if not self.filepath:
            self.entries = []
            self.index = 0
            self.read_flags = []
            self.num_btn.setText('#')
            self._set_num_button_style(is_read=False)
            self.date_label.setText("")
            self.date_label.setToolTip("")
            self.msg_label.setText('### Brak źródła ###')
            self._stop_blink()
            self._update_unread_label()
            self._save_index()
            self._save_read_map()
            self._update_halo_icon()
            return

        try:
            new_entries = self._load_from_spreadsheet(self.filepath) if self._looks_like_spreadsheet(self.filepath) else self._parse_file()
        except Exception:
            new_entries = []

        persisted = self._load_read_map()
        prev_map = {}
        try:
            for idx, e in enumerate(self.entries):
                num = str(e[0])
                if idx < len(self.read_flags):
                    prev_map[num] = 1 if self.read_flags[idx] else 0
        except Exception:
            prev_map = {}

        self.entries = new_entries
        self.read_flags = []
        for e in self.entries:
            num_str = str(e[0]) if e and e[0] is not None else ''
            if num_str in persisted:
                try:
                    flag = bool(int(persisted[num_str]))
                except Exception:
                    flag = bool(persisted[num_str])
            elif num_str in prev_map:
                flag = bool(prev_map[num_str])
            else:
                flag = False
            self.read_flags.append(flag)

        if not self.entries:
            self.index = 0
            self.num_btn.setText('-')
            self._set_num_button_style(is_read=False)
            self.date_label.setText("")
            self.date_label.setToolTip("")
            self.msg_label.setText('Brak komunikatów')
            self._stop_blink()
            self._update_unread_label()
            self._save_index()
            self._save_read_map()
            self._update_halo_icon()
            return

        if self.index >= len(self.entries):
            self.index = 0

        try:
            saved_index = int(self.settings.value(self._index_key, 0))
            if 0 <= saved_index < len(self.entries):
                self.index = saved_index
        except Exception:
            pass

        self.show_current()
        self._update_unread_label()
        self._save_index()
        self._save_read_map()
        self._update_halo_icon()

    def _set_num_button_style(self, is_read: bool):
        name = self.num_btn.objectName() or "halo_num_btn"
        color = "black" if is_read else "red"
        self.num_btn.setStyleSheet(
            f"#{name} {{ color: {color}; background: transparent; border: none; }}"
            f"#{name}:hover {{ background: transparent; }}"
            f"#{name}:focus {{ outline: none; background: transparent; }}"
        )

    def _unread_count(self):
        try:
            return len(self.read_flags) - sum(1 for v in self.read_flags if v)
        except Exception:
            return 0

    def _update_unread_label(self):
        try:
            unread = self._unread_count()
            self.unread_btn.setText(f"Nieprzeczytane: {unread}")
            self.unread_btn.setStyleSheet("color: red;" if unread > 0 else "color: green;")
        except Exception:
            try:
                self.unread_btn.setText("Nieprzeczytane: 0")
                self.unread_btn.setStyleSheet("color: green;")
            except Exception:
                pass

    def _update_halo_icon(self):
        unread = self._unread_count()
        if unread > 0 and self._halo_colored:
            if not self.halo_blink_timer.isActive():
                self.halo_blink_timer.start()
        else:
            if self.halo_blink_timer.isActive():
                self.halo_blink_timer.stop()
            self.halo_icon_label.setPixmap(QPixmap())

    def _start_blink(self):
        if not self.blink_timer.isActive():
            self._blink_state = False
            self.blink_timer.start()

    def _stop_blink(self):
        if self.blink_timer.isActive():
            self.blink_timer.stop()
        try:
            self.btn_down.setStyleSheet("")
        except Exception:
            pass

    def _on_blink_timeout(self):
        self._blink_state = not self._blink_state
        self.btn_down.setStyleSheet("background-color: red; color: white;" if self._blink_state else "")

    def _on_halo_blink(self):
        self._halo_blink_state = not self._halo_blink_state
        if self._halo_blink_state:
            if self._halo_colored:
                self.halo_icon_label.setPixmap(
                    self._halo_colored.scaled(
                        self.halo_icon_label.size(),
                        Qt.KeepAspectRatio,
                        Qt.SmoothTransformation
                    )
                )
            else:
                self.halo_icon_label.setPixmap(QPixmap())
        else:
            self.halo_icon_label.setPixmap(QPixmap())

    def show_current(self):
        if not self.entries:
            try:
                self.num_btn.setText("-")
                self._set_num_button_style(is_read=False)
                self.date_label.setText("")
                self.date_label.setToolTip("")
                self.msg_label.setText("Brak komunikatów")
            except Exception:
                pass
            self._stop_blink()
            self._update_unread_label()
            self._save_index()
            self._update_halo_icon()
            return

        num, dt_obj, raw_date, msg_html = self.entries[self.index]

        self.num_btn.setText(str(num) if num is not None and str(num).strip() else str(self.index + 1))
        try:
            self._set_num_button_style(is_read=bool(self.read_flags[self.index]))
        except Exception:
            self._set_num_button_style(is_read=False)

        short_date = ""
        tooltip_date = ""

        if dt_obj:
            try:
                short_date = f"{dt_obj.day:02d} {self._months_pl[dt_obj.month - 1]} {str(dt_obj.year)[-2:]}"
                tooltip_date = dt_obj.strftime("%d.%m.%Y %H:%M")
            except Exception:
                short_date = ""
                tooltip_date = ""
        elif raw_date:
            parsed = self._try_parse_any_datetime(raw_date)
            if parsed:
                try:
                    short_date = f"{parsed.day:02d} {self._months_pl[parsed.month - 1]} {str(parsed.year)[-2:]}"
                except Exception:
                    short_date = raw_date.split()[0]
                tooltip_date = raw_date
            else:
                token = raw_date.split()[0]
                parsed2 = self._try_parse_any_datetime(token)
                if parsed2:
                    try:
                        short_date = f"{parsed2.day:02d} {self._months_pl[parsed2.month - 1]} {str(parsed2.year)[-2:]}"
                    except Exception:
                        short_date = token
                else:
                    short_date = token
                tooltip_date = raw_date

        self.date_label.setText(short_date)
        self.date_label.setToolTip(tooltip_date)

        try:
            self.msg_label.setText(str(msg_html).replace('\n', '<br>'))
        except Exception:
            self.msg_label.setText(str(msg_html))

        if self.index < len(self.entries) - 1:
            self._start_blink()
        else:
            self._stop_blink()

        self._save_index()
        self._save_read_map()
        self._update_unread_label()
        self._update_halo_icon()

    def next_entry(self):
        if not self.entries:
            return
        self.index = (self.index + 1) % len(self.entries)
        self._last_unread_pos = -1
        self.show_current()

    def prev_entry(self):
        if not self.entries:
            return
        self.index = (self.index - 1) % len(self.entries)
        self._last_unread_pos = -1
        self.show_current()

    def _on_num_clicked(self):
        if not self.entries:
            return
        try:
            self.read_flags[self.index] = not bool(self.read_flags[self.index])
            self._set_num_button_style(is_read=bool(self.read_flags[self.index]))
            self._last_unread_pos = -1
            self._update_unread_label()
            self._save_read_map()
            self._update_halo_icon()
        except Exception:
            pass

    def _on_unread_clicked(self):
        if not self.entries:
            return
        unread_indices = [i for i, v in enumerate(self.read_flags) if not v]
        if not unread_indices:
            QMessageBox.information(self.iface.mainWindow(), "Brak nieprzeczytanych", "Wszystkie komunikaty oznaczone jako przeczytane.")
            return

        next_pos = 0
        if self._last_unread_pos != -1:
            try:
                for j, entry_idx in enumerate(unread_indices):
                    if entry_idx > self._last_unread_pos:
                        next_pos = j
                        break
                else:
                    next_pos = 0
            except Exception:
                next_pos = 0

        target_entry_idx = unread_indices[next_pos]
        self.index = target_entry_idx
        self.show_current()
        self._last_unread_pos = target_entry_idx

    def _on_mark_all_clicked(self):
        if not self.entries:
            return
        try:
            unread = [i for i, v in enumerate(self.read_flags) if not v]
            if unread:
                for i in range(len(self.read_flags)):
                    self.read_flags[i] = True
            else:
                for i in range(len(self.read_flags)):
                    self.read_flags[i] = False
            self._save_read_map()
            self._update_unread_label()
            self._update_halo_icon()
            self.show_current()
        except Exception as e:
            self._log(f"Error in _on_mark_all_clicked: {e}", Qgis.Warning)

    def _on_auto_refresh(self):
        if self._is_refreshing:
            return
        if not self.filepath:
            return
        self._is_refreshing = True
        try:
            self.reload_entries()
        finally:
            self._is_refreshing = False

    def _on_add_clicked(self):
        dlg = QDialog(self.iface.mainWindow())
        dlg.setWindowTitle("Dodaj komunikat")
        v = QVBoxLayout()

        te = QTextEdit()
        te.setPlaceholderText("Wpisz treść komunikatu (wiele wierszy)...")
        te.setMinimumSize(420, 220)
        v.addWidget(te)

        sig = QLineEdit()
        sig.setPlaceholderText("Podpis (zapamiętany)...")
        saved_sig = self.settings.value(self._signature_key, '')
        if saved_sig:
            sig.setText(saved_sig)
        v.addWidget(sig)

        bb = QDialogButtonBox()
        send_btn = bb.addButton("Wyślij 🚀", QDialogButtonBox.AcceptRole)
        cancel_btn = bb.addButton(QDialogButtonBox.Cancel)
        v.addWidget(bb)
        dlg.setLayout(v)

        def on_send():
            message_text = te.toPlainText().strip()
            signature_text = sig.text().strip()
            if not message_text:
                QMessageBox.warning(self.iface.mainWindow(), "Brak treści", "Wprowadź treść komunikatu przed wysłaniem.")
                return

            try:
                if signature_text:
                    self.settings.setValue(self._signature_key, signature_text)
            except Exception:
                pass

            lines = [ln.rstrip() for ln in message_text.splitlines()]
            if signature_text:
                lines.append(f"-- {signature_text}")
            full_msg = "\n".join(lines).strip()

            success, info = self._append_new_entry(full_msg)
            if success:
                QMessageBox.information(self.iface.mainWindow(), "Wysłano", "Komunikat został dodany.")
                te.clear()
                dlg.accept()
                self.reload_entries()
                try:
                    self.index = len(self.entries) - 1
                    self.show_current()
                except Exception:
                    pass
            else:
                QMessageBox.warning(self.iface.mainWindow(), "Niepowodzenie", f"Nie udało się dodać komunikatu.\n{info}")
                self._log(f"Append failed: {info}", Qgis.Warning)

        send_btn.clicked.connect(on_send)
        cancel_btn.clicked.connect(lambda: dlg.reject())
        dlg.exec_()

    def _append_new_entry(self, message_text: str):
        try:
            numeric_vals = []
            for e in self.entries:
                try:
                    numeric_vals.append(int(str(e[0]).strip()))
                except Exception:
                    continue

            new_num = max(numeric_vals) + 1 if numeric_vals else len(self.entries) + 1
            date_str = datetime.now().strftime("%H:%M %d.%m.%Y")

            lines = message_text.splitlines()
            first_line = lines[0] if lines else ""
            remaining = lines[1:] if len(lines) > 1 else []

            block_lines = [f"{new_num};{date_str};{first_line}"]
            block_lines.extend(remaining)
            block_lines.append("***")
            block_text = "\n".join(block_lines) + "\n"

            if self.filepath and os.path.exists(self.filepath):
                try:
                    with open(self.filepath, "ab") as fhb:
                        try:
                            fhb.seek(-1, os.SEEK_END)
                            last = fhb.read(1)
                            if last != b'\n':
                                fhb.write(b'\n')
                        except Exception:
                            pass
                    with open(self.filepath, "a", encoding="utf-8") as fh:
                        fh.write(block_text)
                    self._log(f"Dopisano komunikat lokalnie: {new_num}", Qgis.Info)
                    self.reload_entries()
                    return True, "OK"
                except Exception as e:
                    return False, f"Błąd zapisu lokalnego: {e}"

            webapp_url = str(self.settings.value(self._webapp_url_key, '')).strip()
            webapp_token = str(self.settings.value(self._webapp_token_key, '')).strip()
            if not webapp_url and self.filepath and 'script.google.com' in self.filepath:
                webapp_url = self.filepath
                webapp_token = str(self.settings.value(self._webapp_token_key, '')).strip()

            if webapp_url:
                ok, info = self._post_to_webapp(webapp_url, webapp_token, message_text, new_num)
                if ok:
                    self._log(f"Dopisano komunikat przez WebApp: {new_num}", Qgis.Info)
                    self.reload_entries()
                    return True, "OK"
                return False, f"WebApp error: {info}"

            if self.filepath and (self.filepath.lower().startswith("http://") or self.filepath.lower().startswith("https://")):
                fetched = self._fetch_csv_from_url(self.filepath)
                if fetched is None:
                    return False, "Nie udało się pobrać źródła z URL; nie można dopisać bez dostępu do zapisu."
                kind, data = fetched
                if kind == 'xlsx':
                    return False, "Zdalny plik XLSX nie może być modyfikowany bez API/OAuth."

                try:
                    existing_text = data.decode('utf-8', errors='replace')
                    if not existing_text.endswith("\n"):
                        existing_text += "\n"
                    new_text = existing_text + block_text
                    req = urllib.request.Request(
                        self.filepath,
                        data=new_text.encode('utf-8'),
                        method='PUT',
                        headers={
                            'User-Agent': 'HaloPlugin/1.0',
                            'Content-Type': 'text/plain; charset=utf-8'
                        }
                    )
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        code = resp.getcode()
                        if 200 <= code < 300:
                            self._log(f"Zaktualizowano zdalny plik (PUT) kod={code}", Qgis.Info)
                            self.reload_entries()
                            return True, "OK"
                        return False, f"Serwer zwrócił kod {code} przy próbie zapisu."
                except urllib.error.HTTPError as he:
                    try:
                        body = he.read().decode('utf-8', errors='replace')
                    except Exception:
                        body = str(he)
                    return False, f"HTTPError przy zapisie (PUT): {he.code} {body}"
                except Exception as e:
                    return False, f"Nie udało się zapisać zdalnego pliku (PUT): {e}"

            return False, "Nieznany/nieobsługiwany typ źródła — można dodać tylko do lokalnego pliku lub skonfigurowanego WebApp."
        except Exception as ex:
            return False, f"Wyjątek: {ex}"

    def _post_to_webapp(self, webapp_url, token, text, num=None):
        try:
            payload = {
                "token": token or "",
                "text": text
            }
            if num is not None:
                try:
                    payload["num"] = int(num)
                except Exception:
                    payload["num"] = str(num)

            data = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(
                webapp_url,
                data=data,
                headers={
                    'Content-Type': 'application/json; charset=utf-8',
                    'User-Agent': 'HaloPlugin/1.0'
                },
                method='POST'
            )
            with urllib.request.urlopen(req, timeout=25) as resp:
                code = resp.getcode()
                body = resp.read().decode('utf-8', errors='replace')
                try:
                    j = json.loads(body)
                except Exception:
                    j = {"raw": body}
                if 200 <= code < 300:
                    return True, j
                return False, f"HTTP {code}: {body}"
        except urllib.error.HTTPError as he:
            try:
                body = he.read().decode('utf-8', errors='replace')
            except Exception:
                body = str(he)
            return False, f"HTTPError {he.code}: {body}"
        except Exception as e:
            return False, f"Exception: {e}"

    def _load_halo_pixmap(self, icon_path):
        try:
            icon = QIcon(icon_path)
            pm = icon.pixmap(QSize(48, 48))
            self._halo_colored = pm if pm and not pm.isNull() else None
        except Exception:
            self._halo_colored = None